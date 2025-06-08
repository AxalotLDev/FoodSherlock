import sys
from openpyxl import load_workbook
import numpy as np
from stl import mesh

def read_vertices(sheet):
    verts = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        x, y, z = row
        verts.append([float(x), float(y), float(z)])
    return np.array(verts)

def read_faces(sheet):
    faces = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        i1, i2, i3 = row
        faces.append([int(i1), int(i2), int(i3)])
    return np.array(faces)

def xlsx_to_stl(xlsx_path, stl_path, vert_name=None, face_name=None):
    wb = load_workbook(filename=xlsx_path, data_only=True)
    sheets = wb.sheetnames

    # Если не заданы имена, пробуем автоматически
    if vert_name is None or face_name is None:
        print("Доступные листы в файле:", sheets)
        if len(sheets) < 2:
            raise ValueError("В файле должно быть минимум два листа с вершинами и гранями.")
        vert_name, face_name = sheets[0], sheets[1]
        print(f"Используем автоматически: vertices='{vert_name}', faces='{face_name}'")

    # Проверка наличия листов
    if vert_name not in sheets or face_name not in sheets:
        raise ValueError(
            f"В файле должны быть листы '{vert_name}' и '{face_name}'. "
            f"Найдены: {sheets}"
        )

    verts = read_vertices(wb[vert_name])
    faces = read_faces(wb[face_name])

    your_mesh = mesh.Mesh(np.zeros(faces.shape[0], dtype=mesh.Mesh.dtype))
    for i, f in enumerate(faces):
        for j in range(3):
            your_mesh.vectors[i][j] = verts[f[j]]

    your_mesh.save(stl_path)
    print(f"Готово! Сохранено: {stl_path}")

if __name__ == "__main__":
    # Возможные варианты запуска:
    # 1) python converter.py input.xlsx output.stl
    # 2) python converter.py input.xlsx output.stl vertices faces
    argc = len(sys.argv)
    if argc not in (3, 5):
        print(
            "Использование:\n"
            "  python converter.py input.xlsx output.stl\n"
            "  python converter.py input.xlsx output.stl vertices faces"
        )
        sys.exit(1)

    xlsx_path = sys.argv[1]
    stl_path = sys.argv[2]
    vert_name = sys.argv[3] if argc == 5 else None
    face_name = sys.argv[4] if argc == 5 else None

    xlsx_to_stl(xlsx_path, stl_path, vert_name, face_name)
